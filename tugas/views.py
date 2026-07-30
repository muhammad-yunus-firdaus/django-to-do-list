import json

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.exceptions import ValidationError
from datetime import date, timedelta, time as dt_time, datetime as dt_datetime

from .models import Tugas, Subtask, AktivitasHarian, EvaluasiMingguan, Kegiatan
from .forms import (
    TugasForm, RegisterForm, CustomAuthenticationForm, SubtaskForm,
    ProfilForm, GantiPasswordForm, AktivitasHarianForm, EvaluasiForm,
    KegiatanForm,
)
from .notifications import check_subtask_completion, generate_deadline_notifications, generate_overdue_notifications
from .services import (
    get_dashboard_stats,
    get_upcoming_deadlines,
    get_high_priority_tasks,
    get_approaching_deadline_tasks,
    get_overdue_tasks,
    get_filtered_tugas,
    mark_task_complete,
    get_export_data,
)


# ══════════════════════════════════════════════════════════════════════
# AUTH - Login, Logout, Register
# ══════════════════════════════════════════════════════════════════════

def login_view(request):
    # Halaman login, auto redirect ke dashboard kalo udah login
    if request.user.is_authenticated:
        return redirect("tugas:dashboard")

    if request.method == "POST":
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Login berhasil! Selamat datang, {user.username}.")
            return redirect(request.GET.get("next") or "tugas:dashboard")
        else:
            messages.error(request, "Username atau password salah.")
    else:
        form = CustomAuthenticationForm(request)

    return render(request, "tugas/login.html", {"form": form})


@login_required
def logout_view(request):
    # Logout user terus balik ke halaman login
    logout(request)
    messages.info(request, "Kamu telah logout.")
    return redirect("tugas:login")


def register(request):
    # Halaman daftar akun baru, cuma pake username doang tanpa email
    if request.user.is_authenticated:
        return redirect("tugas:dashboard")

    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Registrasi berhasil! Silakan login terlebih dahulu.")
            return redirect("tugas:login")
    else:
        form = RegisterForm()

    return render(request, "tugas/register.html", {"form": form})


# ══════════════════════════════════════════════════════════════════════
# DASHBOARD - Statistik & Overview
# ══════════════════════════════════════════════════════════════════════

@login_required
def dashboard_view(request):
    # Halaman utama setelah login, nampilin semua statistik tugas
    user = request.user

    # Ambil semua data statistik dari service layer
    stats = get_dashboard_stats(user)
    tugas_deadline_terdekat = get_upcoming_deadlines(user)
    prioritas_tinggi = get_high_priority_tasks(user)
    tugas_mendekati_deadline = get_approaching_deadline_tasks(user)
    tugas_overdue = get_overdue_tasks(user)

    # Generate notifikasi otomatis waktu buka dashboard
    generate_deadline_notifications()
    generate_overdue_notifications()

    # NOTE: Django messages udah ga dipake lagi, sekarang pake notification system
    # if tugas_mendekati_deadline.exists():
    #     messages.warning(
    #         request,
    #         f"Ada {tugas_mendekati_deadline.count()} tugas yang mendekati deadline (kurang dari 2 hari)!"
    #     )

    # if tugas_overdue.exists():
    #     messages.error(
    #         request,
    #         f"Ada {tugas_overdue.count()} tugas yang sudah melewati deadline!"
    #     )

    # Ambil agenda hari ini untuk widget ringkasan jadwal
    agenda_hari_ini = AktivitasHarian.objects.filter(
        user=user,
        tanggal=date.today(),
    ).select_related('user').order_by('jam_mulai')

    agenda_total = agenda_hari_ini.count()
    agenda_selesai = agenda_hari_ini.filter(status='selesai').count()
    agenda_persen = round((agenda_selesai / agenda_total) * 100, 1) if agenda_total > 0 else 0

    # Ambil kegiatan mendatang (maksimal 5)
    kegiatan_mendatang = Kegiatan.objects.filter(
        user=user,
        status='akan_datang',
        tanggal__gte=date.today(),
    ).select_related('user').order_by('tanggal', 'jam_mulai')[:5]

    # Hitung kegiatan hari ini untuk statistik gabungan
    kegiatan_hari_ini = Kegiatan.objects.filter(
        user=user,
        tanggal=date.today(),
    )
    kegiatan_total_today = kegiatan_hari_ini.count()
    kegiatan_selesai_today = kegiatan_hari_ini.filter(status='selesai').count()

    # Hitung progres keseluruhan gabungan (Tugas + Aktivitas Hari Ini + Kegiatan Hari Ini)
    total_tugas = stats.get("total_tugas", 0)
    tugas_selesai = stats.get("tugas_selesai", 0)
    total_items = total_tugas + agenda_total + kegiatan_total_today
    completed_items = tugas_selesai + agenda_selesai + kegiatan_selesai_today
    belum_items = total_items - completed_items
    progres_persen = round((completed_items / total_items * 100)) if total_items > 0 else 0

    context = {
        **stats,
        "tugas_deadline_terdekat": tugas_deadline_terdekat,
        "prioritas_tinggi": prioritas_tinggi,
        "tugas_mendekati_deadline": tugas_mendekati_deadline,
        "tugas_overdue": tugas_overdue,
        "agenda_hari_ini": agenda_hari_ini,
        "agenda_total": agenda_total,
        "agenda_selesai": agenda_selesai,
        "agenda_persen": agenda_persen,
        "kegiatan_mendatang": kegiatan_mendatang,
        "total_items": total_items,
        "completed_items": completed_items,
        "belum_items": belum_items,
        "progres_persen": progres_persen,
    }

    return render(request, "tugas/dashboard.html", context)


# ══════════════════════════════════════════════════════════════════════
# CRUD TUGAS - Create, Read, Update, Delete
# ══════════════════════════════════════════════════════════════════════

@login_required
def daftar_tugas(request):
    # Nampilin semua tugas punya user ini, bisa di-filter dan di-search
    filters = {
        "prioritas": request.GET.get("prioritas", ""),
        "kategori": request.GET.get("kategori", ""),
        "status": request.GET.get("status", ""),
        "q": request.GET.get("q", ""),
    }

    tugas_list = get_filtered_tugas(request.user, filters)
    kategori_list = [k[0] for k in Tugas.KATEGORI_CHOICES]

    # Buat pagination biar ga terlalu banyak di satu halaman
    paginator = Paginator(tugas_list, 10)
    page_number = request.GET.get('page')
    
    try:
        tugas_page = paginator.page(page_number)
    except PageNotAnInteger:
        tugas_page = paginator.page(1)
    except EmptyPage:
        tugas_page = paginator.page(paginator.num_pages)

    context = {
        "tugas_list": tugas_page,
        "kategori_list": kategori_list,
        "prioritas_filter": filters["prioritas"],
        "kategori_filter": filters["kategori"],
        "status_filter": filters["status"],
        "search_query": filters["q"],
    }
    return render(request, "tugas/daftar_tugas.html", context)


@login_required
def tambah_tugas(request):
    # Form buat bikin tugas baru + dynamic subtask support
    if request.method == "POST":
        form = TugasForm(request.POST)
        if form.is_valid():
            tugas = form.save(commit=False)
            tugas.user = request.user
            tugas.save()

            # Tangkap subtask dinamis dari form
            subtask_list = request.POST.getlist('subtasks[]')
            for idx, judul in enumerate(subtask_list):
                judul_clean = judul.strip()
                if judul_clean:  # Skip subtask kosong
                    Subtask.objects.create(
                        tugas=tugas,
                        judul=judul_clean,
                        urutan=idx + 1,
                    )

            messages.success(request, "Tugas berhasil ditambahkan!")
            return redirect("tugas:daftar")
    else:
        form = TugasForm()

    return render(request, "tugas/tambah_tugas.html", {"form": form})


@login_required
def edit_tugas(request, tugas_id):
    # Form buat edit tugas yang udah ada
    tugas = get_object_or_404(Tugas, id=tugas_id, user=request.user)

    if request.method == "POST":
        form = TugasForm(request.POST, instance=tugas)
        if form.is_valid():
            updated_tugas = form.save(commit=False)

            if not form.cleaned_data.get("deadline"):
                updated_tugas.deadline = tugas.deadline

            updated_tugas.save()
            messages.success(request, "Tugas berhasil diperbarui!")
            return redirect("tugas:daftar")
        else:
            messages.error(request, "Terjadi kesalahan. Periksa kembali form kamu.")
    else:
        form = TugasForm(instance=tugas)

    return render(request, "tugas/edit_tugas.html", {"form": form})


@login_required
@require_POST
def hapus_tugas(request, tugas_id):
    # Hapus tugas, cuma bisa lewat POST biar aman
    tugas = get_object_or_404(Tugas, id=tugas_id, user=request.user)
    tugas.delete()
    messages.success(request, "Tugas berhasil dihapus!")
    return redirect("tugas:daftar")


@login_required
@require_POST
def tandai_selesai(request, tugas_id):
    # Tandain tugas jadi selesai, cuma bisa lewat POST
    tugas = get_object_or_404(Tugas, id=tugas_id, user=request.user)

    if mark_task_complete(tugas):
        messages.success(request, f"Tugas '{tugas.judul}' telah selesai!")
    else:
        messages.warning(request, "Tugas ini sudah selesai sebelumnya!")

    return redirect("tugas:daftar")


@login_required
def detail_tugas(request, tugas_id):
    # Halaman detail tugas, nampilin info lengkap sama subtask-nya
    tugas = get_object_or_404(Tugas, id=tugas_id, user=request.user)
    return render(request, "tugas/detail_tugas.html", {"tugas": tugas})


# ══════════════════════════════════════════════════════════════════════
# EXPORT - CSV, Excel, PDF
# ══════════════════════════════════════════════════════════════════════

@login_required
def export_csv(request):
    # Download semua tugas dalam format CSV
    import csv

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="tugas.csv"'

    writer = csv.writer(response)
    writer.writerow(['Judul', 'Kategori', 'Prioritas', 'Deadline', 'Status'])

    for row in get_export_data(request.user):
        writer.writerow([
            row["judul"], row["kategori"], row["prioritas"],
            row["deadline"], row["status"]
        ])

    return response


@login_required
def export_excel(request):
    # Download tugas dalam format Excel (.xlsx) dengan styling
    from openpyxl import Workbook
    from openpyxl.styles.borders import Border, Side
    from openpyxl.styles import Font, Alignment, PatternFill

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="tugas.xlsx"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daftar Tugas"

    # Kasih styling biar lebih rapi
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Judul", "Kategori", "Prioritas", "Deadline", "Status"]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

    for row_idx, row in enumerate(get_export_data(request.user), start=2):
        sheet.append([
            row["judul"], row["kategori"], row["prioritas"],
            row["deadline"], row["status"]
        ])
        for cell in sheet[row_idx]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border_style

    # Atur lebar kolom otomatis sesuai isi
    for col in sheet.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        sheet.column_dimensions[col[0].column_letter].width = max_len + 2

    workbook.save(response)
    return response


@login_required
def export_pdf(request):
    # Download tugas dalam format PDF dengan tabel rapi
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="tugas.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]

    data = [["Judul", "Deskripsi", "Kategori", "Prioritas", "Deadline", "Status"]]

    for row in get_export_data(request.user):
        data.append([
            Paragraph(row["judul"], style_normal),
            Paragraph(row["deskripsi"], style_normal),
            row["kategori"], row["prioritas"],
            row["deadline"], row["status"],
        ])

    table = Table(data, colWidths=[150, 250, 100, 80, 120, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
    ]))

    doc.build([table])
    return response


@login_required
def export_kegiatan_pdf(request):
    """Download all kegiatan/acara in PDF format."""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="kegiatan.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]

    data = [["Judul", "Deskripsi", "Tanggal", "Jam Mulai", "Jam Selesai", "Lokasi", "Kategori", "Status"]]

    kegiatan = Kegiatan.objects.filter(user=request.user).order_by('tanggal', 'jam_mulai')
    for keg in kegiatan:
        data.append([
            Paragraph(keg.judul, style_normal),
            Paragraph(keg.deskripsi or '-', style_normal),
            keg.tanggal.isoformat(),
            keg.jam_mulai.strftime('%H:%M'),
            keg.jam_selesai.strftime('%H:%M') if keg.jam_selesai else '-',
            Paragraph(keg.lokasi or '-', style_normal),
            keg.get_kategori_display(),
            keg.get_status_display(),
        ])

    table = Table(data, colWidths=[120, 180, 80, 60, 60, 100, 100, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
    ]))

    doc.build([table])
    return response


@login_required
def export_kegiatan_excel(request):
    """Download all kegiatan/acara in Excel (.xlsx) format."""
    from openpyxl import Workbook
    from openpyxl.styles.borders import Border, Side
    from openpyxl.styles import Font, Alignment, PatternFill

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="kegiatan.xlsx"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kegiatan & Acara"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Judul", "Deskripsi", "Tanggal", "Jam Mulai", "Jam Selesai", "Lokasi", "Kategori", "Status"]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

    kegiatan = Kegiatan.objects.filter(user=request.user).order_by('tanggal', 'jam_mulai')
    for row_idx, keg in enumerate(kegiatan, start=2):
        sheet.append([
            keg.judul,
            keg.deskripsi or '-',
            keg.tanggal.isoformat(),
            keg.jam_mulai.strftime('%H:%M'),
            keg.jam_selesai.strftime('%H:%M') if keg.jam_selesai else '-',
            keg.lokasi or '-',
            keg.get_kategori_display(),
            keg.get_status_display()
        ])
        for cell in sheet[row_idx]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border_style

    for col in sheet.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        sheet.column_dimensions[col[0].column_letter].width = max_len + 2

    workbook.save(response)
    return response


# ══════════════════════════════════════════════════════════════════════
# SUBTASK - Kelola Sub-tugas
# ══════════════════════════════════════════════════════════════════════

@login_required
@require_POST
def tambah_subtask(request, tugas_id):
    # Tambahin subtask baru ke dalam tugas
    tugas = get_object_or_404(Tugas, id=tugas_id, user=request.user)
    form = SubtaskForm(request.POST)
    
    if form.is_valid():
        subtask = form.save(commit=False)
        subtask.tugas = tugas
        
        # Kasih nomor urutan otomatis (ambil yang terakhir + 1)
        last_subtask = tugas.subtasks.order_by('-urutan').first()
        subtask.urutan = (last_subtask.urutan + 1) if last_subtask else 1
        
        subtask.save()
        messages.success(request, f"Subtask '{subtask.judul}' berhasil ditambahkan!")
    else:
        messages.error(request, "Gagal menambahkan subtask. Silakan coba lagi.")
    
    return redirect('tugas:detail', tugas_id=tugas.id)


@login_required
@require_POST
def edit_subtask(request, subtask_id):
    # Ubah judul subtask
    subtask = get_object_or_404(Subtask, id=subtask_id, tugas__user=request.user)
    
    judul_baru = request.POST.get('judul', '').strip()
    if judul_baru:
        subtask.judul = judul_baru
        subtask.save()
        messages.success(request, "Subtask berhasil diupdate!")
    else:
        messages.error(request, "Judul subtask tidak boleh kosong.")
    
    return redirect('tugas:detail', tugas_id=subtask.tugas.id)


@login_required
@require_POST
def hapus_subtask(request, subtask_id):
    # Hapus subtask dari tugas
    subtask = get_object_or_404(Subtask, id=subtask_id, tugas__user=request.user)
    tugas_id = subtask.tugas.id
    judul = subtask.judul
    
    subtask.delete()
    messages.success(request, f"Subtask '{judul}' berhasil dihapus!")
    
    return redirect('tugas:detail', tugas_id=tugas_id)


@login_required
@require_POST
def toggle_subtask(request, subtask_id):
    # Toggle centang/uncentang subtask (dipanggil via AJAX)
    subtask = get_object_or_404(Subtask, id=subtask_id, tugas__user=request.user)
    
    subtask.selesai = not subtask.selesai
    subtask.save()
    
    # Cek kalo semua subtask udah selesai, bikin notifikasi
    check_subtask_completion(subtask.tugas)
    
    # Kirim balik response JSON buat update UI
    return JsonResponse({
        'success': True,
        'selesai': subtask.selesai,
        'progress': subtask.tugas.subtask_progress
    })


# ══════════════════════════════════════════════════════════════════════
# PROFIL - Edit Username, Email, Password
# ══════════════════════════════════════════════════════════════════════

@login_required
def edit_profil_view(request):
    """Halaman edit profil: ubah username/email dan ganti password."""
    user = request.user
    profil_form = ProfilForm(user=user)
    password_form = GantiPasswordForm(user=user)
    profil_success = False
    password_success = False

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "update_profil":
            profil_form = ProfilForm(request.POST, user=user)
            if profil_form.is_valid():
                profil_form.save()
                messages.success(request, "Profil berhasil diperbarui!")
                profil_success = True
                return redirect("tugas:edit_profil")
            else:
                messages.error(request, "Gagal memperbarui profil. Periksa kembali data kamu.")

        elif action == "update_password":
            password_form = GantiPasswordForm(request.POST, user=user)
            if password_form.is_valid():
                password_form.save()
                # Update session supaya ga auto-logout setelah ganti password
                update_session_auth_hash(request, user)
                messages.success(request, "Password berhasil diubah!")
                password_success = True
                return redirect("tugas:edit_profil")
            else:
                messages.error(request, "Gagal mengubah password. Periksa kembali data kamu.")

    context = {
        "profil_form": profil_form,
        "password_form": password_form,
        "profil_success": profil_success,
        "password_success": password_success,
    }
    return render(request, "tugas/edit_profil.html", context)


# ══════════════════════════════════════════════════════════════════════
# AGENDA HARIAN - Daily Time-Blocking + Habit Tracker
# ══════════════════════════════════════════════════════════════════════

def _copy_habits_for_today(user):
    """
    Salin semua aktivitas habit ke tanggal hari ini jika belum ada.
    Dipanggil on-demand saat halaman agenda dibuka (PythonAnywhere compatible).
    """
    today = date.today()

    # Ambil semua habit unik milik user (dari hari manapun)
    habits = AktivitasHarian.objects.filter(
        user=user,
        is_habit=True,
    ).values('judul', 'jam_mulai', 'durasi_menit').distinct()

    for habit in habits:
        # Cek apakah sudah ada aktivitas dengan judul + jam yang sama di hari ini
        exists = AktivitasHarian.objects.filter(
            user=user,
            tanggal=today,
            judul=habit['judul'],
            jam_mulai=habit['jam_mulai'],
        ).exists()

        if not exists:
            try:
                AktivitasHarian.objects.create(
                    user=user,
                    judul=habit['judul'],
                    jam_mulai=habit['jam_mulai'],
                    durasi_menit=habit['durasi_menit'],
                    is_habit=True,
                    tanggal=today,
                    status='belum',
                )
            except ValidationError:
                # Jika overlap dengan aktivitas lain, skip tanpa error
                pass


def _get_next_free_slot(user, tanggal):
    """
    Hitung slot waktu kosong berikutnya berdasarkan aktivitas/kegiatan terakhir di hari tersebut.
    - Jika ada, ambil jam_selesai terakhir + 1 menit.
    - Jika belum ada dan tanggal = hari ini, default ke waktu sekarang (dibulatkan ke 5 menit berikutnya).
    - Jika belum ada dan bukan hari ini, default ke 07:00.
    """
    last_activity = AktivitasHarian.objects.filter(
        user=user,
        tanggal=tanggal,
    ).order_by('-jam_mulai').first()

    last_kegiatan = Kegiatan.objects.filter(
        user=user,
        tanggal=tanggal,
    ).order_by('-jam_mulai').first()

    latest_time = None
    if last_activity and last_activity.jam_selesai:
        latest_time = last_activity.jam_selesai
    if last_kegiatan and last_kegiatan.jam_selesai:
        if latest_time is None or last_kegiatan.jam_selesai > latest_time:
            latest_time = last_kegiatan.jam_selesai

    if latest_time:
        # Tambah 1 menit dari jam selesai terakhir
        last_end_dt = dt_datetime.combine(tanggal, latest_time)
        next_slot_dt = last_end_dt + timedelta(minutes=1)
        # Pastikan tidak melewati 23:59
        if next_slot_dt.time() < dt_time(23, 59):
            return next_slot_dt.time().strftime('%H:%M')
        return '23:00'

    # Belum ada aktivitas/kegiatan
    if tanggal == date.today():
        from django.utils.timezone import localtime
        now_local = localtime()
        # Bulatkan ke 5 menit berikutnya
        minute = now_local.minute
        rounded_minute = ((minute // 5) + 1) * 5
        if rounded_minute >= 60:
            next_hour = now_local.hour + 1
            if next_hour > 23:
                return '23:00'
            return dt_time(next_hour, 0).strftime('%H:%M')
        return dt_time(now_local.hour, rounded_minute).strftime('%H:%M')

    return '07:00'


def _generate_24h_timeline(aktivitas_list, kegiatan_list):
    """
    Membuat timeline 24 jam dengan slot 30 menit.
    Aktivitas/kegiatan disisipkan secara presisi ke slot yang sesuai.
    Slot kosong ditandai sebagai 'kosong'.
    Returns list of dicts siap dijadikan JSON.
    """
    # Kumpulkan semua occupied intervals
    occupied = []
    for akt in aktivitas_list:
        end_time = akt.jam_selesai or akt._calculate_jam_selesai()
        if end_time:
            occupied.append({
                'type': 'aktivitas',
                'id': akt.id,
                'judul': akt.judul,
                'jam_mulai': akt.jam_mulai.strftime('%H:%M'),
                'jam_selesai': end_time.strftime('%H:%M'),
                'durasi_menit': akt.durasi_menit,
                'status': akt.status,
                'is_habit': akt.is_habit,
                'start_minutes': akt.jam_mulai.hour * 60 + akt.jam_mulai.minute,
                'end_minutes': end_time.hour * 60 + end_time.minute,
            })
    for keg in kegiatan_list:
        if keg.jam_selesai:
            occupied.append({
                'type': 'kegiatan',
                'id': keg.id,
                'judul': keg.judul,
                'jam_mulai': keg.jam_mulai.strftime('%H:%M'),
                'jam_selesai': keg.jam_selesai.strftime('%H:%M'),
                'kategori': keg.kategori,
                'status': keg.status,
                'lokasi': keg.lokasi or '',
                'start_minutes': keg.jam_mulai.hour * 60 + keg.jam_mulai.minute,
                'end_minutes': keg.jam_selesai.hour * 60 + keg.jam_selesai.minute,
            })

    # Sort by start time
    occupied.sort(key=lambda x: x['start_minutes'])

    # Build timeline: iterate through 24h in 30-min slots, inserting items precisely
    timeline = []
    cursor = 0  # current position in minutes from 00:00

    for item in occupied:
        # Fill gap before this item with empty slots (30-min chunks)
        while cursor < item['start_minutes']:
            slot_end = min(cursor + 30, item['start_minutes'])
            timeline.append({
                'type': 'kosong',
                'jam_mulai': f"{cursor // 60:02d}:{cursor % 60:02d}",
                'jam_selesai': f"{slot_end // 60:02d}:{slot_end % 60:02d}",
            })
            cursor = slot_end

        # Add the occupied item
        timeline.append(item)
        cursor = max(cursor, item['end_minutes'])

    # Fill remaining time after last item until 24:00
    while cursor < 1440:
        slot_end = min(cursor + 30, 1440)
        end_h = slot_end // 60 if slot_end < 1440 else 23
        end_m = slot_end % 60 if slot_end < 1440 else 59
        timeline.append({
            'type': 'kosong',
            'jam_mulai': f"{cursor // 60:02d}:{cursor % 60:02d}",
            'jam_selesai': f"{end_h:02d}:{end_m:02d}",
        })
        cursor = slot_end

    return timeline


@login_required
def agenda_harian_view(request):
    """Halaman agenda harian: timeline terpadu aktivitas + kegiatan + form tambah + navigasi tanggal."""
    # Ambil tanggal dari query param, default hari ini
    tanggal_str = request.GET.get('tanggal', '')
    if tanggal_str:
        try:
            tanggal = date.fromisoformat(tanggal_str)
        except ValueError:
            tanggal = date.today()
    else:
        tanggal = date.today()

    # Copy habits untuk hari ini (on-demand)
    if tanggal == date.today():
        _copy_habits_for_today(request.user)

    # Ambil semua aktivitas di tanggal tersebut, dioptimasi dengan select_related
    aktivitas_list = AktivitasHarian.objects.filter(
        user=request.user,
        tanggal=tanggal,
    ).select_related('user').order_by('jam_mulai')

    # Ambil semua kegiatan di tanggal tersebut
    kegiatan_list = Kegiatan.objects.filter(
        user=request.user,
        tanggal=tanggal,
    ).select_related('user').order_by('jam_mulai')

    # Gabungkan aktivitas dan kegiatan menjadi timeline terpadu
    timeline_items = []
    for akt in aktivitas_list:
        timeline_items.append({
            'type': 'aktivitas',
            'item': akt,
            'jam_mulai': akt.jam_mulai,
        })
    for keg in kegiatan_list:
        timeline_items.append({
            'type': 'kegiatan',
            'item': keg,
            'jam_mulai': keg.jam_mulai,
        })
    timeline_items.sort(key=lambda x: x['jam_mulai'])

    # Hitung statistik (berdasarkan aktivitas harian)
    total = aktivitas_list.count()
    selesai = aktivitas_list.filter(status='selesai').count()
    persen = round((selesai / total) * 100, 1) if total > 0 else 0

    # ── Preserve Form Input saat Overlap Error ──
    form_has_error = False
    saved_form_data = request.session.pop('agenda_form_data', None)

    if saved_form_data:
        # Kembalikan data form dari session (setelah overlap error)
        form = AktivitasHarianForm(initial=saved_form_data)
        form_has_error = True
    else:
        # ── Smart Default Jam Mulai ──
        next_slot = _get_next_free_slot(request.user, tanggal)
        form = AktivitasHarianForm(initial={'jam_mulai': next_slot})

    # ── Rekomendasi Jadwal Kemarin ──
    show_yesterday_prompt = False
    yesterday_activities = []
    if total == 0:
        kemarin = tanggal - timedelta(days=1)
        akt_kemarin = AktivitasHarian.objects.filter(
            user=request.user,
            tanggal=kemarin,
        ).select_related('user').order_by('jam_mulai')
        if akt_kemarin.exists():
            show_yesterday_prompt = True
            yesterday_activities = list(akt_kemarin.values(
                'judul', 'jam_mulai', 'durasi_menit', 'is_habit'
            ))
            # Format jam_mulai ke string untuk JSON serialization di template
            for akt in yesterday_activities:
                akt['jam_mulai'] = akt['jam_mulai'].strftime('%H:%M')

    # Generate timeline 24 jam
    timeline_24h_json = _generate_24h_timeline(aktivitas_list, kegiatan_list)

    # Navigasi tanggal
    prev_date = tanggal - timedelta(days=1)
    next_date = tanggal + timedelta(days=1)

    context = {
        "aktivitas_list": aktivitas_list,
        "timeline_items": timeline_items,
        "timeline_24h_json": timeline_24h_json,
        "tanggal": tanggal,
        "is_today": tanggal == date.today(),
        "prev_date": prev_date,
        "next_date": next_date,
        "form": form,
        "form_has_error": form_has_error,
        "total_aktivitas": total,
        "aktivitas_selesai": selesai,
        "persen_aktivitas": persen,
        "show_yesterday_prompt": show_yesterday_prompt,
        "yesterday_activities_json": yesterday_activities,
    }
    return render(request, "tugas/agenda_harian.html", context)


@login_required
def tambah_aktivitas_view(request):
    """Tambah aktivitas harian baru dengan validasi anti-overlap."""
    if request.method == "POST":
        form = AktivitasHarianForm(request.POST)
        tanggal_str = request.POST.get('tanggal', '')
        tanggal = date.today()
        if tanggal_str:
            try:
                tanggal = date.fromisoformat(tanggal_str)
            except ValueError:
                pass

        # Set user and tanggal on form instance before validation so clean() check works
        form.instance.user = request.user
        form.instance.tanggal = tanggal

        # Helper: simpan form data ke session untuk preserve saat error
        def _save_form_to_session():
            request.session['agenda_form_data'] = {
                'judul': request.POST.get('judul', ''),
                'jam_mulai': request.POST.get('jam_mulai', ''),
                'durasi_menit': request.POST.get('durasi_menit', ''),
                'is_habit': 'is_habit' in request.POST,
            }

        if form.is_valid():
            try:
                form.save()
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f"Aktivitas '{form.instance.judul}' berhasil ditambahkan!"
                    })
                messages.success(request, f"Aktivitas '{form.instance.judul}' berhasil ditambahkan!")
            except ValidationError as e:
                error_msg = "; ".join(e.messages) if hasattr(e, 'messages') else str(e)
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': error_msg
                    }, status=400)
                messages.error(request, error_msg)
                _save_form_to_session()
        else:
            errors_list = []
            for field, errors in form.errors.items():
                for error in errors:
                    prefix = "" if field == "__all__" else f"{field.capitalize()}: "
                    errors_list.append(f"{prefix}{error}")
            error_msg = "; ".join(errors_list)
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=400)
            # Display form errors to user
            for field, errors in form.errors.items():
                for error in errors:
                    prefix = "" if field == "__all__" else f"{field.capitalize()}: "
                    messages.error(request, f"{prefix}{error}")
            _save_form_to_session()

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        return redirect(f"/tugas/agenda/?tanggal={tanggal.isoformat()}")

    return redirect("tugas:agenda")


@login_required
@require_POST
def toggle_aktivitas_view(request, aktivitas_id):
    """Toggle atau ubah status aktivitas (AJAX-friendly)."""
    import json
    aktivitas = get_object_or_404(AktivitasHarian, id=aktivitas_id, user=request.user)
    
    target_status = None
    try:
        data = json.loads(request.body)
        target_status = data.get('status')
    except (json.JSONDecodeError, ValueError, AttributeError):
        pass

    if target_status in ['belum', 'selesai', 'terlewat']:
        if aktivitas.status == 'selesai' and target_status == 'terlewat':
            return JsonResponse({
                'success': False,
                'error': 'Aktivitas yang sudah selesai tidak boleh ditandai terlewat.'
            }, status=400)
        aktivitas.status = target_status
    else:
        # Fallback toggle biasa (selesai <-> belum)
        if aktivitas.status == 'selesai':
            aktivitas.status = 'belum'
        else:
            aktivitas.status = 'selesai'

    # Save tanpa full_clean untuk menghindari overlap check saat toggle status
    aktivitas.save(update_fields=['status', 'updated_at'])

    # Hitung ulang statistik
    tanggal = aktivitas.tanggal
    all_akt = AktivitasHarian.objects.filter(user=request.user, tanggal=tanggal)
    total = all_akt.count()
    selesai = all_akt.filter(status='selesai').count()
    persen = round((selesai / total) * 100, 1) if total > 0 else 0

    return JsonResponse({
        'success': True,
        'status': aktivitas.status,
        'total': total,
        'selesai': selesai,
        'persen': persen,
    })


@login_required
def edit_aktivitas_view(request, aktivitas_id):
    """Edit data aktivitas harian."""
    aktivitas = get_object_or_404(AktivitasHarian, id=aktivitas_id, user=request.user)
    if request.method == "POST":
        form = AktivitasHarianForm(request.POST, instance=aktivitas)
        
        # Set user dan tanggal sebelum validasi
        form.instance.user = request.user
        form.instance.tanggal = aktivitas.tanggal

        if form.is_valid():
            try:
                form.save()
                messages.success(request, f"Aktivitas '{form.instance.judul}' berhasil diperbarui!")
            except ValidationError as e:
                error_msg = "; ".join(e.messages) if hasattr(e, 'messages') else str(e)
                messages.error(request, error_msg)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    prefix = "" if field == "__all__" else f"{field.capitalize()}: "
                    messages.error(request, f"{prefix}{error}")
                    
        return redirect(f"/tugas/agenda/?tanggal={aktivitas.tanggal.isoformat()}")

    return redirect("tugas:agenda")


@login_required
@require_POST
def hapus_aktivitas_view(request, aktivitas_id):
    """Hapus aktivitas harian."""
    aktivitas = get_object_or_404(AktivitasHarian, id=aktivitas_id, user=request.user)
    tanggal = aktivitas.tanggal
    judul = aktivitas.judul
    aktivitas.delete()
    messages.success(request, f"Aktivitas '{judul}' berhasil dihapus!")
    return redirect(f"/tugas/agenda/?tanggal={tanggal.isoformat()}")


@login_required
def api_jadwal_hari_ini(request):
    """API endpoint: return jadwal hari ini berstatus 'belum' untuk notifikasi pengingat."""
    today = date.today()
    aktivitas = AktivitasHarian.objects.filter(
        user=request.user,
        tanggal=today,
        status='belum',
    ).order_by('jam_mulai')

    kegiatan = Kegiatan.objects.filter(
        user=request.user,
        tanggal=today,
        status='akan_datang',
    ).order_by('jam_mulai')

    items = []
    for akt in aktivitas:
        items.append({
            'judul': akt.judul,
            'jam_mulai': akt.jam_mulai.strftime('%H:%M'),
            'type': 'aktivitas',
        })
    for keg in kegiatan:
        items.append({
            'judul': keg.judul,
            'jam_mulai': keg.jam_mulai.strftime('%H:%M'),
            'type': 'kegiatan',
        })

    items.sort(key=lambda x: x['jam_mulai'])
    return JsonResponse({'items': items})


@login_required
@require_POST
def copy_jadwal_kemarin_view(request):
    """Copy jadwal dari hari kemarin (atau tanggal tertentu) ke tanggal target."""
    import json
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Data tidak valid.'}, status=400)

    tanggal_str = data.get('tanggal', date.today().isoformat())
    try:
        tanggal_target = date.fromisoformat(tanggal_str)
    except ValueError:
        tanggal_target = date.today()

    aktivitas_items = data.get('aktivitas', [])
    created_count = 0
    skipped_count = 0

    for item in aktivitas_items:
        judul = item.get('judul', '').strip()
        jam_mulai_str = item.get('jam_mulai', '')
        durasi_menit = item.get('durasi_menit', 30)
        is_habit = item.get('is_habit', False)

        if not judul or not jam_mulai_str:
            continue

        try:
            jam_mulai = dt_time.fromisoformat(jam_mulai_str)
        except ValueError:
            continue

        try:
            durasi_int = int(durasi_menit)
        except (ValueError, TypeError):
            durasi_int = 30

        try:
            AktivitasHarian.objects.create(
                user=request.user,
                judul=judul,
                jam_mulai=jam_mulai,
                durasi_menit=durasi_int,
                is_habit=is_habit,
                tanggal=tanggal_target,
                status='belum',
            )
            created_count += 1
        except ValidationError:
            skipped_count += 1

    return JsonResponse({
        'success': True,
        'created': created_count,
        'skipped': skipped_count,
        'message': f'{created_count} aktivitas berhasil disalin.'
                   + (f' {skipped_count} dilewati karena jadwal bertabrakan.' if skipped_count > 0 else ''),
    })
# ══════════════════════════════════════════════════════════════════════
# EVALUASI MINGGUAN - Ringkasan Statistik + Catatan Refleksi
# ══════════════════════════════════════════════════════════════════════

def _get_week_range(ref_date=None):
    """Hitung tanggal awal (Senin) dan akhir (Minggu) dari minggu yang mengandung ref_date."""
    if ref_date is None:
        ref_date = date.today()
    # weekday(): Monday=0, Sunday=6
    start = ref_date - timedelta(days=ref_date.weekday())
    end = start + timedelta(days=6)
    return start, end


def _get_weekly_stats(user, start_date, end_date):
    """
    Hitung statistik mingguan secara real-time.
    Arsitektur terisolasi — siap untuk diintegrasikan dengan Gemini AI.
    """
    # Statistik Tugas Utama
    from django.utils.timezone import make_aware
    from datetime import datetime as dt

    tugas_qs = Tugas.objects.filter(user=user)
    # Tugas yang dibuat dalam rentang minggu tersebut
    total_tugas = tugas_qs.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).count()
    tugas_selesai = tugas_qs.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status='selesai',
    ).count()
    persen_tugas = round((tugas_selesai / total_tugas) * 100, 1) if total_tugas > 0 else 0

    # Statistik Aktivitas Harian
    akt_qs = AktivitasHarian.objects.filter(
        user=user,
        tanggal__gte=start_date,
        tanggal__lte=end_date,
    )
    total_aktivitas = akt_qs.count()
    aktivitas_selesai = akt_qs.filter(status='selesai').count()
    persen_aktivitas = round((aktivitas_selesai / total_aktivitas) * 100, 1) if total_aktivitas > 0 else 0

    return {
        "total_tugas": total_tugas,
        "tugas_selesai": tugas_selesai,
        "persen_tugas": persen_tugas,
        "total_aktivitas": total_aktivitas,
        "aktivitas_selesai": aktivitas_selesai,
        "persen_aktivitas": persen_aktivitas,
    }


@login_required
def evaluasi_view(request):
    """Halaman evaluasi mingguan: ringkasan statistik + catatan refleksi + riwayat."""
    user = request.user

    # Hitung rentang minggu ini
    minggu_mulai, minggu_selesai = _get_week_range()

    # Hitung statistik real-time
    stats = _get_weekly_stats(user, minggu_mulai, minggu_selesai)

    # Cek apakah sudah ada evaluasi untuk minggu ini
    evaluasi_existing = EvaluasiMingguan.objects.filter(
        user=user,
        minggu_mulai=minggu_mulai,
    ).first()

    if request.method == "POST":
        form = EvaluasiForm(request.POST)
        if form.is_valid():
            catatan = form.cleaned_data.get("catatan_evaluasi", "")

            if evaluasi_existing:
                # Update evaluasi yang sudah ada
                evaluasi_existing.total_tugas = stats["total_tugas"]
                evaluasi_existing.tugas_selesai = stats["tugas_selesai"]
                evaluasi_existing.persen_tugas = stats["persen_tugas"]
                evaluasi_existing.total_aktivitas = stats["total_aktivitas"]
                evaluasi_existing.aktivitas_selesai = stats["aktivitas_selesai"]
                evaluasi_existing.persen_aktivitas = stats["persen_aktivitas"]
                evaluasi_existing.catatan_evaluasi = catatan
                evaluasi_existing.save()
            else:
                # Buat evaluasi baru
                EvaluasiMingguan.objects.create(
                    user=user,
                    minggu_mulai=minggu_mulai,
                    minggu_selesai=minggu_selesai,
                    total_tugas=stats["total_tugas"],
                    tugas_selesai=stats["tugas_selesai"],
                    persen_tugas=stats["persen_tugas"],
                    total_aktivitas=stats["total_aktivitas"],
                    aktivitas_selesai=stats["aktivitas_selesai"],
                    persen_aktivitas=stats["persen_aktivitas"],
                    catatan_evaluasi=catatan,
                )

            messages.success(request, "Evaluasi mingguan berhasil disimpan!")
            return redirect("tugas:evaluasi")
    else:
        initial_catatan = evaluasi_existing.catatan_evaluasi if evaluasi_existing else ""
        form = EvaluasiForm(initial={"catatan_evaluasi": initial_catatan})

    # Ambil riwayat evaluasi sebelumnya, dioptimasi dengan select_related
    evaluasi_history = EvaluasiMingguan.objects.filter(user=request.user).select_related('user').order_by('-created_at')

    # Ambil rincian data untuk minggu ini, dioptimasi dengan select_related dan prefetch_related
    tugas_minggu_ini = Tugas.objects.filter(
        user=user,
        deadline__date__range=[minggu_mulai, minggu_selesai]
    ).select_related('user').prefetch_related('subtasks').order_by('deadline')
    
    aktivitas_minggu_ini = AktivitasHarian.objects.filter(
        user=user,
        tanggal__range=[minggu_mulai, minggu_selesai]
    ).select_related('user').order_by('tanggal', 'jam_mulai')

    context = {
        "minggu_mulai": minggu_mulai,
        "minggu_selesai": minggu_selesai,
        "stats": stats,
        "form": form,
        "evaluasi_existing": evaluasi_existing,
        "evaluasi_history": evaluasi_history,
        "tugas_minggu_ini": tugas_minggu_ini,
        "aktivitas_minggu_ini": aktivitas_minggu_ini,
    }
    return render(request, "tugas/evaluasi.html", context)


# ══════════════════════════════════════════════════════════════════════
# KEGIATAN & ACARA - CRUD & Views
# ══════════════════════════════════════════════════════════════════════

@login_required
def kegiatan_list_view(request):
    """Halaman utama daftar & form kegiatan acara dengan filter."""
    kategori = request.GET.get('kategori', '')
    status_filter = request.GET.get('status_filter', 'akan_datang')
    
    qs = Kegiatan.objects.filter(user=request.user).select_related('user')
    
    if kategori:
        qs = qs.filter(kategori=kategori)
        
    today = date.today()
    if status_filter == 'akan_datang':
        qs = qs.filter(status='akan_datang').order_by('tanggal', 'jam_mulai')
    elif status_filter == 'riwayat':
        from django.db.models import Q
        qs = qs.filter(Q(status__in=['selesai', 'dibatalkan']) | Q(tanggal__lt=today)).order_by('-tanggal', '-jam_mulai')
    else:
        qs = qs.order_by('tanggal', 'jam_mulai')

    form = KegiatanForm()
    
    context = {
        'kegiatan_list': qs,
        'form': form,
        'kategori_filter': kategori,
        'status_filter': status_filter,
        'kategori_choices': Kegiatan.KATEGORI_CHOICES,
        'today': today,
    }
    return render(request, 'tugas/kegiatan_list.html', context)


@login_required
@require_POST
def kegiatan_tambah_view(request):
    """Tambah kegiatan acara baru via AJAX."""
    form = KegiatanForm(request.POST)
    if form.is_valid():
        kegiatan = form.save(commit=False)
        kegiatan.user = request.user
        try:
            kegiatan.save()
            return JsonResponse({'success': True})
        except ValidationError as e:
            return JsonResponse({'success': False, 'error': e.messages[0] if hasattr(e, 'messages') else str(e)})
    else:
        errors = []
        for field, err_list in form.errors.items():
            for err in err_list:
                errors.append(f"{form.fields[field].label or field}: {err}")
        return JsonResponse({'success': False, 'error': "; ".join(errors)})


@login_required
def kegiatan_edit_view(request, kegiatan_id):
    """Edit kegiatan acara via AJAX."""
    kegiatan = get_object_or_404(Kegiatan, id=kegiatan_id, user=request.user)
    if request.method == "POST":
        form = KegiatanForm(request.POST, instance=kegiatan)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({'success': True})
            except ValidationError as e:
                return JsonResponse({'success': False, 'error': e.messages[0] if hasattr(e, 'messages') else str(e)})
        else:
            errors = []
            for field, err_list in form.errors.items():
                for err in err_list:
                    errors.append(f"{form.fields[field].label or field}: {err}")
            return JsonResponse({'success': False, 'error': "; ".join(errors)})
    else:
        # Request data kegiatan dalam bentuk JSON
        data = {
            'id': kegiatan.id,
            'judul': kegiatan.judul,
            'kategori': kegiatan.kategori,
            'tanggal': kegiatan.tanggal.isoformat(),
            'jam_mulai': kegiatan.jam_mulai.strftime('%H:%M'),
            'jam_selesai': kegiatan.jam_selesai.strftime('%H:%M'),
            'lokasi': kegiatan.lokasi,
            'catatan': kegiatan.catatan,
            'status': kegiatan.status,
        }
        return JsonResponse({'success': True, 'data': data})


@login_required
@require_POST
def kegiatan_hapus_view(request, kegiatan_id):
    """Hapus kegiatan acara via AJAX."""
    kegiatan = get_object_or_404(Kegiatan, id=kegiatan_id, user=request.user)
    kegiatan.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def kegiatan_toggle_status_view(request, kegiatan_id):
    """Toggle atau ubah status kegiatan langsung (AJAX-friendly)."""
    import json
    kegiatan = get_object_or_404(Kegiatan, id=kegiatan_id, user=request.user)
    
    target_status = None
    try:
        data = json.loads(request.body)
        target_status = data.get('status')
    except (json.JSONDecodeError, ValueError, AttributeError):
        pass

    if target_status in ['akan_datang', 'selesai', 'dibatalkan']:
        kegiatan.status = target_status
        try:
            kegiatan.save(update_fields=['status', 'updated_at'])
            return JsonResponse({'success': True, 'status': kegiatan.status})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    else:
        return JsonResponse({'success': False, 'error': 'Status tidak valid.'}, status=400)



@login_required
def export_jadwal_pdf_view(request):
    """Download daily agenda schedule in PDF format."""
    from datetime import date
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    # Ambil tanggal
    tanggal_str = request.GET.get('tanggal', '')
    if tanggal_str:
        try:
            tanggal = date.fromisoformat(tanggal_str)
        except ValueError:
            tanggal = date.today()
    else:
        tanggal = date.today()

    user = request.user
    aktivitas_list = AktivitasHarian.objects.filter(
        user=user,
        tanggal=tanggal,
    ).select_related('user').order_by('jam_mulai')

    kegiatan_list = Kegiatan.objects.filter(
        user=user,
        tanggal=tanggal,
    ).select_related('user').order_by('jam_mulai')

    timeline_items = []
    for akt in aktivitas_list:
        timeline_items.append({
            'type': 'Aktivitas Harian',
            'judul': akt.judul,
            'jam_mulai': akt.jam_mulai,
            'jam_selesai': akt.jam_selesai,
            'kategori': 'Habit' if akt.is_habit else 'Rutin',
            'durasi': f"{akt.durasi_menit} mnt",
            'lokasi': '-',
            'status': akt.get_status_display() if hasattr(akt, 'get_status_display') else akt.status,
        })
        
    for keg in kegiatan_list:
        timeline_items.append({
            'type': 'Kegiatan & Acara',
            'judul': keg.judul,
            'jam_mulai': keg.jam_mulai,
            'jam_selesai': keg.jam_selesai,
            'kategori': keg.get_kategori_display() if hasattr(keg, 'get_kategori_display') else keg.kategori,
            'durasi': '-',
            'lokasi': keg.lokasi or '-',
            'status': keg.get_status_display() if hasattr(keg, 'get_status_display') else keg.status,
        })

    timeline_items.sort(key=lambda x: x['jam_mulai'])

    # Hitung statistik
    total_items = len(timeline_items)
    selesai_items = sum(1 for item in timeline_items if item['status'].lower() in ['selesai'])
    progres_persen = round((selesai_items / total_items) * 100) if total_items > 0 else 0

    response = HttpResponse(content_type="application/pdf")
    filename = f"jadwal_{tanggal.isoformat()}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Build PDF doc A4 portrait
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=54, leftMargin=54, topMargin=54, bottomMargin=54)
    styles = getSampleStyleSheet()

    # Kustom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        'DocMeta',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#475569'),
        spaceAfter=15,
    )
    summary_style = ParagraphStyle(
        'SummaryText',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=20,
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#1E293B'),
    )
    header_cell_style = ParagraphStyle(
        'HeaderCellText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
    )

    story = []

    # Title
    story.append(Paragraph(f"Jadwal & Agenda Harian - {user.username}", title_style))
    # Tanggal
    formatted_date = tanggal.strftime('%d %B %Y')
    story.append(Paragraph(f"Tanggal: {formatted_date}", meta_style))

    # Summary
    summary_text = (
        f"<b>Ringkasan Hari Ini:</b> Total {total_items} item agenda | "
        f"{selesai_items} Selesai | Progres: {progres_persen}%"
    )
    story.append(Paragraph(summary_text, summary_style))

    # Table headers
    headers = [
        Paragraph("Jam", header_cell_style),
        Paragraph("Agenda / Kegiatan", header_cell_style),
        Paragraph("Kategori", header_cell_style),
        Paragraph("Durasi", header_cell_style),
        Paragraph("Lokasi / Link", header_cell_style),
        Paragraph("Status", header_cell_style)
    ]
    data = [headers]

    for item in timeline_items:
        jam_str = f"{item['jam_mulai'].strftime('%H:%M')} - {item['jam_selesai'].strftime('%H:%M') if item['jam_selesai'] else '?'}"
        data.append([
            Paragraph(jam_str, cell_style),
            Paragraph(f"<b>{item['judul']}</b><br/><font color='#64748B' size='8'>{item['type']}</font>", cell_style),
            Paragraph(item['kategori'], cell_style),
            Paragraph(item['durasi'], cell_style),
            Paragraph(item['lokasi'], cell_style),
            Paragraph(f"<b>{item['status']}</b>", cell_style)
        ])

    # 487 pt total width
    table = Table(data, colWidths=[80, 147, 75, 45, 90, 50])
    
    # Alternating rows style
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#4338CA')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
    ]
    
    for i in range(1, len(data)):
        if i % 2 == 0:
            t_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
            
    table.setStyle(TableStyle(t_style))
    story.append(table)

    doc.build(story)
    return response